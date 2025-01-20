from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class saveToExcel:
    def __init__(self, excelSavePath, excel, title):
        self.excelSavePath = excelSavePath          # excel的保存路径
        self.excel = excel                       # openpyxl.Workbook()的实例话
        self.sheet = self.excel.create_sheet(title=title)   # 创建工作区
        self.Sheet_line = 1               # 表格的行

    def save_list_to_excel(self, title, list_result):
        self.sheet.cell(self.Sheet_line, 1).value = title
        self.Sheet_line += 1
        for each_result in list_result:
            self.sheet.cell(self.Sheet_line, 1).value = each_result
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    def save_dict_to_excel(self, list_dict_result):
        try:
            for j, key in enumerate(list_dict_result[0].keys()):
                self.sheet.cell(self.Sheet_line, j+1).value = key
            self.Sheet_line += 1
            for each_result in list_dict_result:
                for j, (key, value) in enumerate(each_result.items()):
                    self.sheet.cell(self.Sheet_line, j+1).value = value
                self.Sheet_line += 1
            self.excel.save(self.excelSavePath)
        except Exception as e:
            pass






